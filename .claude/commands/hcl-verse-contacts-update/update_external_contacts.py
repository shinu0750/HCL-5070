#!/usr/bin/env python3
"""
讀回 hcl-verse-RAG 產生的 external_contacts.xlsx（人工填好 canonical_name 的列），
回填三個地方：
  1. email_mapping（PostgreSQL，upsert，不會清掉手動加的列）
  2. Qdrant verse_emails collection（set_payload 更新 from_name）
  3. Hindsight EID bank（先 get_document 讀回舊 tags/metadata，合併新姓名後
     重新 retain——retain() 是整段覆蓋不是 merge，沒帶到的 tags/metadata 會消失）
最後把 external_contacts_state.json 對應的聯絡人標成 confirmed=true，
下次 hcl-verse-RAG 產生 Excel 時就不會再列出來。

用法：
    python update_external_contacts.py [xlsx_path]
"""
import os, sys, json, warnings
warnings.filterwarnings('ignore')

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import psycopg2
import requests
from openpyxl import load_workbook
from qdrant_client import QdrantClient
from qdrant_client.models import Filter, FieldCondition, MatchValue

_env_path = os.path.expanduser("~/.hermes/.env")
if os.path.exists(_env_path):
    with open(_env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, _, v = line.partition('=')
                os.environ.setdefault(k.strip(), v.strip())

# 共用 hcl-verse-RAG 既有的 tracker 模組（state 檔案讀寫），不重複實作
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "hcl-verse-RAG"))
from external_contacts_tracker import load_state, save_state

QDRANT_URL   = os.environ.get("QDRANT_URL", "http://10.11.1.40:6333")
HINDSIGHT_URL = os.environ.get("HINDSIGHT_URL", "http://localhost:8888/mcp/")
COLLECTION   = "verse_emails"
XLSX_PATH    = os.path.expanduser("~/verse-export/external_contacts.xlsx")

PG_HOST = os.environ.get("PG_HOST", "")
PG_PORT = os.environ.get("PG_PORT", "5432")
PG_DB   = os.environ.get("PG_DB", "")
PG_USER = os.environ.get("PG_USER", "")
PG_PASSWORD = os.environ.get("PG_PASSWORD", "")


class HindsightClient:
    def __init__(self, url):
        self.url = url
        resp = requests.post(url, json={
            "jsonrpc": "2.0", "id": 1, "method": "initialize",
            "params": {"protocolVersion": "2024-11-05", "capabilities": {},
                       "clientInfo": {"name": "update-external-contacts", "version": "1.0"}},
        })
        self.session_id = resp.headers.get("mcp-session-id")

    def _call(self, name, arguments):
        resp = requests.post(self.url, json={
            "jsonrpc": "2.0", "id": 2, "method": "tools/call",
            "params": {"name": name, "arguments": arguments},
        }, headers={"mcp-session-id": self.session_id}, timeout=30)
        for line in resp.text.split("\n"):
            if line.startswith("data:"):
                return json.loads(line[5:])
        return {}

    def get_document(self, document_id, bank_id="EID"):
        """讀回既有記錄的 tags/document_metadata，重新 retain 前一定要先呼叫這個，
        否則 retain() 整段覆蓋會把沒帶到的 tags/metadata 洗掉。"""
        result = self._call("get_document", {"document_id": document_id, "bank_id": bank_id})
        text = result.get("result", {}).get("content", [{}])[0].get("text", "")
        try:
            return json.loads(text)
        except Exception:
            return None

    def retain(self, content, document_id, timestamp, metadata, context, tags=None, bank_id="EID"):
        return self._call("retain", {
            "content": content, "document_id": document_id, "timestamp": timestamp,
            "metadata": metadata, "context": context, "tags": tags, "bank_id": bank_id,
        })


def upsert_email_mapping(email, name):
    conn = psycopg2.connect(host=PG_HOST, port=PG_PORT, dbname=PG_DB, user=PG_USER, password=PG_PASSWORD)
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO email_mapping (name, email) VALUES (%s, %s) "
                "ON CONFLICT (email) DO UPDATE SET name = EXCLUDED.name",
                (name, email),
            )
        conn.commit()
    finally:
        conn.close()


def read_confirmed_rows(xlsx_path):
    """讀 Excel，只取 canonical_name 有填的列，回傳 (workbook, worksheet,
    [(row_idx, email, canonical_name), ...])。呼叫端處理完後用 row_idx 把該列從
    worksheet 刪掉再存檔，避免同一份 Excel 沒重新產生就重跑時被重複處理。"""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    email_idx = header.index("email")
    name_idx = header.index("canonical_name")
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        email = row[email_idx] if email_idx < len(row) else None
        name = row[name_idx] if name_idx < len(row) else None
        if email and name and str(name).strip():
            rows.append((row_idx, str(email).strip().lower(), str(name).strip()))
    return wb, ws, rows


def backfill_one(qdrant, hindsight, email, canonical_name):
    """回填單一聯絡人在 Qdrant/Hindsight 裡的所有記錄，回傳實際更新的筆數。"""
    points, _ = qdrant.scroll(
        collection_name=COLLECTION,
        scroll_filter=Filter(must=[FieldCondition(key="from_email", match=MatchValue(value=email))]),
        limit=1000, with_payload=True,
    )
    if not points:
        print(f"  ↷ Qdrant 查無 {email} 相關資料，跳過（可能當初 RAG 那步失敗過，不是錯誤）")
        return 0

    updated = 0
    for p in points:
        unid = p.payload.get("unid")
        if not unid:
            continue

        qdrant.set_payload(collection_name=COLLECTION, payload={"from_name": canonical_name}, points=[p.id])

        doc = hindsight.get_document(unid)
        if doc is None:
            print(f"  ✗ Hindsight 查無 unid={unid}，跳過（可能當初 retain 那步失敗過，不是錯誤）")
            continue

        old_tags = doc.get("tags") or ["mail"]
        old_metadata = dict(doc.get("document_metadata") or {})
        old_metadata["from_name"] = canonical_name
        old_metadata["from_email"] = email

        subject = old_metadata.get("subject", p.payload.get("subject", ""))
        sent_date = old_metadata.get("sent_date", p.payload.get("sent_date", ""))
        body = p.payload.get("body", "")

        content = (
            f"主旨：{subject}\n"
            f"寄件者：{canonical_name} <{email}>\n"
            f"日期：{sent_date}\n\n{body}"
        )
        result = hindsight.retain(
            content=content, document_id=unid, timestamp=sent_date,
            metadata=old_metadata, tags=old_tags,
            context=f"HCL Verse 信件：主旨「{subject}」，寄件者 {canonical_name}",
        )
        result_text = result.get("result", {}).get("content", [{}])[0].get("text", "")
        if "validation error" in result_text.lower():
            print(f"  ✗ Hindsight retain 被拒絕（unid={unid}）：{result_text[:200]}")
            continue
        updated += 1

    return updated


def _replace_name_in_list(raw, old_names, new_name):
    """to/cc 存的是 resolve_recipients() 組完的『、』分隔姓名字串，不是結構化的
    email 清單，沒有 to_email/cc_email 可以查，只能整段名字完全比對 old_names
    （這個聯絡人被確認前 Verse 顯示過的所有舊名字，來自 external_contacts_state.json
    的 seen_names）才替換，避免誤傷剛好同名的其他字串片段。回傳 (新字串, 是否有變動)。"""
    if not raw:
        return raw, False
    parts = raw.split("、")
    changed = False
    result = []
    for part in parts:
        if part.strip() in old_names:
            result.append(new_name)
            changed = True
        else:
            result.append(part)
    return "、".join(result), changed


def backfill_to_cc(qdrant, hindsight, email, canonical_name, seen_names):
    """把 to/cc 欄位裡這個聯絡人被確認前的舊顯示名（seen_names）換成 canonical_name。

    這個聯絡人可能只出現在別人信件的收件人清單裡（不是 from_email），Qdrant 沒有
    to_email/cc_email 這種結構化欄位可以查，只能全表掃一次比對文字——已經靠
    「處理完的列從 Excel 刪掉」擋住同一個聯絡人被重複處理，這裡不用再另外防重複。
    Hindsight 的 content 本來就不含收件人資訊（只有主旨/寄件者/日期/內文），
    metadata 也沒有 cc（3.7.0 移除），所以只要更新 metadata.to，不用重組 content。
    回傳實際更新的筆數。"""
    old_names = {n.strip() for n in (seen_names or []) if n and n.strip()}
    if not old_names:
        return 0

    updated = 0
    next_offset = None
    while True:
        points, next_offset = qdrant.scroll(
            collection_name=COLLECTION, limit=200, with_payload=True, offset=next_offset,
        )
        for p in points:
            new_to, to_changed = _replace_name_in_list(p.payload.get("to", ""), old_names, canonical_name)
            new_cc, cc_changed = _replace_name_in_list(p.payload.get("cc", ""), old_names, canonical_name)
            if not to_changed and not cc_changed:
                continue

            payload_update = {}
            if to_changed:
                payload_update["to"] = new_to
            if cc_changed:
                payload_update["cc"] = new_cc
            qdrant.set_payload(collection_name=COLLECTION, payload=payload_update, points=[p.id])

            unid = p.payload.get("unid")
            if unid and to_changed:
                doc = hindsight.get_document(unid)
                if doc is None:
                    print(f"  ✗ Hindsight 查無 unid={unid}（to 欄位），跳過（不是錯誤）")
                else:
                    old_tags = doc.get("tags") or ["mail"]
                    old_metadata = dict(doc.get("document_metadata") or {})
                    old_metadata["to"] = new_to
                    result = hindsight.retain(
                        content=doc.get("original_text", ""),
                        document_id=unid,
                        timestamp=old_metadata.get("sent_date", ""),
                        metadata=old_metadata, tags=old_tags,
                        context=f"HCL Verse 信件：主旨「{old_metadata.get('subject', '')}」，"
                                f"寄件者 {old_metadata.get('from_name', '')}",
                    )
                    result_text = result.get("result", {}).get("content", [{}])[0].get("text", "")
                    if "validation error" in result_text.lower():
                        print(f"  ✗ Hindsight to 欄位更新被拒絕（unid={unid}）：{result_text[:200]}")

            updated += 1

        if next_offset is None:
            break

    return updated


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else XLSX_PATH
    if not os.path.exists(xlsx_path):
        print(f"找不到 {xlsx_path}")
        return

    wb, ws, rows = read_confirmed_rows(xlsx_path)
    if not rows:
        print("Excel 裡沒有已填 canonical_name 的列，沒有要回填的")
        return

    qdrant = QdrantClient(url=QDRANT_URL)
    hindsight = HindsightClient(HINDSIGHT_URL)
    state = load_state()

    summary = []
    processed_row_indices = []
    for row_idx, email, canonical_name in rows:
        print(f"處理 {email} -> {canonical_name}")
        seen_names = state.get(email, {}).get("seen_names", [])
        upsert_email_mapping(email, canonical_name)
        n = backfill_one(qdrant, hindsight, email, canonical_name)
        n_to_cc = backfill_to_cc(qdrant, hindsight, email, canonical_name, seen_names)
        if email in state:
            state[email]["confirmed"] = True
        summary.append({
            "email": email, "canonical_name": canonical_name,
            "unids_updated": n, "to_cc_updated": n_to_cc,
        })
        processed_row_indices.append(row_idx)
        print(f"  ✓ 回填 {n} 筆（from_email）+ {n_to_cc} 筆（to/cc）")

    save_state(state)

    # 處理完的列直接從 Excel 刪掉——email_mapping 已經 upsert 進去了，這列留著沒有
    # 意義，留著反而會在同一份 Excel 沒重新產生就重跑時被重複處理一次。由後往前刪，
    # 避免刪除時後面列的 row_idx 跟著往前移
    for row_idx in sorted(processed_row_indices, reverse=True):
        ws.delete_rows(row_idx)
    wb.save(xlsx_path)

    print("\n=== 完成 ===")
    for s in summary:
        print(f"  {s['email']} -> {s['canonical_name']}："
              f"{s['unids_updated']} 筆（from_email）+ {s['to_cc_updated']} 筆（to/cc）")


if __name__ == "__main__":
    main()
