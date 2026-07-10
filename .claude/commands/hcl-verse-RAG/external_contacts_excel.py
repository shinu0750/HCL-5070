#!/usr/bin/env python3
"""
獨立模組：把 external_contacts_tracker 的 state 產生成 Excel，給人填 canonical_name。

重點：重新產生時會先讀舊 Excel，保留使用者已經填好但還沒處理的 canonical_name，
不會覆蓋掉使用者的編輯進度。只把 confirmed=false 的列寫進去（已確認的就不用再顯示）。
"""
import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

COLUMNS = ["email", "seen_names", "count", "first_seen", "last_seen", "canonical_name"]


def read_existing_canonical_names(xlsx_path):
    """讀舊 Excel，只取出已經填的 canonical_name（email -> name），
    合併時不覆蓋使用者已填但還沒處理的內容。"""
    if not os.path.exists(xlsx_path):
        return {}
    wb = load_workbook(xlsx_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    if "email" not in header or "canonical_name" not in header:
        return {}
    email_idx = header.index("email")
    name_idx = header.index("canonical_name")
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if email_idx >= len(row):
            continue
        email = row[email_idx]
        canonical = row[name_idx] if name_idx < len(row) else None
        if email and canonical:
            result[str(email).strip().lower()] = str(canonical).strip()
    return result


def generate_excel(state, xlsx_path):
    """
    state: external_contacts_tracker.load_state() 的結果。
    只寫 confirmed=false 的列；canonical_name 欄位會合併舊 Excel 裡已填的值。
    回傳實際寫入的列數。
    """
    existing_names = read_existing_canonical_names(xlsx_path)
    pending = {email: info for email, info in state.items() if not info.get("confirmed")}

    wb = Workbook()
    ws = wb.active
    ws.title = "待確認聯絡人"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for email in sorted(pending.keys()):
        info = pending[email]
        canonical = existing_names.get(email, "")
        ws.append([
            email,
            "; ".join(info.get("seen_names", [])),
            info.get("count", 0),
            info.get("first_seen", ""),
            info.get("last_seen", ""),
            canonical,
        ])

    widths = {"A": 28, "B": 30, "C": 8, "D": 18, "E": 18, "F": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out_dir = os.path.dirname(xlsx_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(xlsx_path)
    return len(pending)


def _run_self_test():
    import tempfile
    tmp_path = os.path.join(tempfile.mkdtemp(), "test_external_contacts.xlsx")

    # 第一輪：2 個待確認聯絡人
    state = {
        "kanek@cicorp.com.tw": {
            "seen_names": ["kane", "康文胜/Kane"], "count": 3,
            "first_seen": "2026-07-07", "last_seen": "2026-07-08", "confirmed": False,
        },
        "sam@vendor.com": {
            "seen_names": ["Sam Lee"], "count": 1,
            "first_seen": "2026-07-09", "last_seen": "2026-07-09", "confirmed": False,
        },
    }
    n = generate_excel(state, tmp_path)
    print(f"第一輪產生：{n} 列")
    assert n == 2

    wb = load_workbook(tmp_path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print("內容：")
    for r in rows:
        print(" ", r)
    assert len(rows) == 2

    # 模擬使用者填了 kane 的 canonical_name，但還沒填 sam 的
    wb2 = load_workbook(tmp_path)
    ws2 = wb2.active
    for row in ws2.iter_rows(min_row=2):
        if row[0].value == "kanek@cicorp.com.tw":
            row[5].value = "康文胜/Kane"
    wb2.save(tmp_path)
    print("\n模擬使用者填好 kane 的 canonical_name 後存檔")

    # 第二輪：sam 又出現一次（count 增加），且多了一個新聯絡人 lisa
    state["sam@vendor.com"]["count"] = 2
    state["sam@vendor.com"]["last_seen"] = "2026-07-10"
    state["lisa@newvendor.com"] = {
        "seen_names": ["Lisa Wang"], "count": 1,
        "first_seen": "2026-07-10", "last_seen": "2026-07-10", "confirmed": False,
    }
    n2 = generate_excel(state, tmp_path)
    print(f"\n第二輪重新產生：{n2} 列")
    assert n2 == 3

    wb3 = load_workbook(tmp_path)
    ws3 = wb3.active
    rows3 = {r[0]: r for r in ws3.iter_rows(min_row=2, values_only=True)}
    print("合併後內容：")
    for email, r in rows3.items():
        print(" ", r)

    assert rows3["kanek@cicorp.com.tw"][5] == "康文胜/Kane", "使用者已填的 canonical_name 應該被保留"
    assert rows3["sam@vendor.com"][5] in (None, ""), "sam 還沒填，應該還是空的"
    assert rows3["sam@vendor.com"][2] == 2, "sam 的 count 應該更新成 2"
    assert "lisa@newvendor.com" in rows3, "新聯絡人 lisa 應該被加進去"

    # 第三輪：kane 確認完畢（confirmed=True），下次產生時應該從清單消失
    state["kanek@cicorp.com.tw"]["confirmed"] = True
    n3 = generate_excel(state, tmp_path)
    wb4 = load_workbook(tmp_path)
    ws4 = wb4.active
    emails4 = [r[0] for r in ws4.iter_rows(min_row=2, values_only=True)]
    print(f"\n第三輪（kane 已確認）：{n3} 列 -> {emails4}")
    assert "kanek@cicorp.com.tw" not in emails4, "已確認的 kane 應該從清單消失"

    print("\n全部自我測試通過 ✓")


if __name__ == "__main__":
    _run_self_test()
